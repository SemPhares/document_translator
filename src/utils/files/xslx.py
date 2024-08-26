import openpyxl
import os
from src.utils.log import logger


def read_xlsx_file(file):
    document = Document(file)
    
    # split the document into pages

    text = []
    for paragraph in document.paragraphs:
        text.append(paragraph.text)
    return '\n'.join(text)


def split_xlsx(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Create a directory to save the split sheets
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_dir = f"{base_name}_sheets"
    os.makedirs(output_dir, exist_ok=True)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Create a new workbook for the current sheet
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name
        
        # Copy the sheet content to the new workbook
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)
        
        # Save the new workbook
        sheet_file_name = os.path.join(output_dir, f"{base_name}_{sheet_name}.xlsx")
        new_workbook.save(sheet_file_name)
        print(f"Saved: {sheet_file_name}")


def extract_images_from_doc(pdf_path):
    images = convert_from_path(pdf_path)
    image_files = []
    for i, image in enumerate(images):
        image_file = f'image_{i + 1}.png'
        image.save(image_file, 'PNG')
        image_files.append(image_file)
    return image_files



def save_translated_xlsx(translated_pages,
                         original_pdf):
    
    output_pdf = io.BytesIO()
    pdf_writer = PyPDF2.PdfWriter()
    
    for page_num, translated_text in sorted(translated_pages.items()):
        page = original_pdf.pages[page_num]
        pdf_writer.add_blank_page(width=page.mediabox.width,
                                height=page.mediabox.width)
        # Ajouter le texte traduit à la page PDF
        pdf_writer.add_page 
        # pages[page_num].merge_page(translated_text, expand = True)

    pdf_writer.write(output_pdf)
    output_pdf.seek(0)
    return output_pdf
