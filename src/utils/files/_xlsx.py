import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

def save_image(image:OpenpyxlImage, filename):
    """Enregistre l'image dans un fichier."""
    with open(filename, 'wb') as img_file:
        img_file.write(image._data())


def extract_images_from_excel(file_path, output_dir):
    """Extrait toutes les images d'un fichier Excel et les enregistre."""
    # Chargement du fichier Excel
    wb = load_workbook(file_path, data_only=True)

    # Création du répertoire de sortie s'il n'existe pas
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Parcours de toutes les feuilles du classeur
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"Extraction des images de la feuille: {sheet}")
        
        # Parcours de tous les objets de la feuille
        for image in ws._images:
            if isinstance(image, OpenpyxlImage):
                # Nom du fichier pour l'image
                img_name = f"{sheet}_image_{ws._images.index(image) + 1}.png"
                img_path = os.path.join(output_dir, img_name)
                
                # Sauvegarde de l'image
                save_image(image, img_path)
                print(f"Image sauvegardée: {img_path}")



def save_sheets_as_text(file_path, output_dir):
    # Lire le fichier Excel et charger les feuilles
    xls = pd.ExcelFile(file_path)
    
    # Créer le répertoire de sortie s'il n'existe pas
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    for sheet_name in xls.sheet_names:
        # Lire la feuille en DataFrame
        df = pd.read_excel(xls, sheet_name=sheet_name).dropna(how='all', axis=1).dropna(how='all', axis=0)
        print(df.to_string(index=False))
        
        # Convertir la DataFrame en texte
        sheet_text = df.to_string(index=False)
        
        # Nom du fichier texte de sortie
        output_file = os.path.join(output_dir, f"{sheet_name}.txt")
        
        # Enregistrer le texte dans un fichier
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(sheet_text)
        
        print(f"Feuille {sheet_name} sauvegardée dans {output_file}")



if __name__ == "__main__":
    file_path = "/Users/elsem/Documents/code/document_translator/test/doc/perception-belbin-feuille.xlsx"

    output_dir = 'output_images'      
    # extract_images_from_excel(file_path, output_dir)
    save_sheets_as_text(file_path, output_dir)